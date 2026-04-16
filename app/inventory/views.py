import csv
import re
import openpyxl
from django.conf import settings
from django.core.files.storage import default_storage
from django.forms import inlineformset_factory
from app.inventory.forms import (
    ProductForm,
    InventoryImageForm,
    InventoryImageUploadForm,
    InventoryForm,
    CategoryForm,
)
from app.inventory.models import Inventory, InventoryImage, Category
from app.storefront.models import StorefrontProductImage
from app.core.s3_uploader import upload_fileobj, S3UploadError
from django.contrib.auth.decorators import login_required
from app.core.decorators import role_required
from django.http import HttpResponse, JsonResponse
from decimal import Decimal
from decimal import InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from .models import Product, Inventory
from django.db import connection, reset_queries, transaction
from django.db.models import F, Q
from django.db.models.functions import Lower
from app.core.tenant_middleware import get_current_tenant


def _wants_json(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in request.headers.get('accept', '')


def _normalize_header(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value).strip().lower())


def _parse_decimal(value):
    cleaned = _normalize_number(value)
    if cleaned is None:
        return None
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_int(value):
    cleaned = _normalize_number(value)
    if cleaned is None:
        return None
    try:
        return int(Decimal(cleaned))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _normalize_number(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith('(') and text.endswith(')'):
        text = f"-{text[1:-1]}"
    text = re.sub(r'[\s\u00A0]', '', text)
    text = text.replace(',', '')
    text = re.sub(r'[^0-9.\-]', '', text)
    return text or None


def _clean_str(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value).strip())


def _public_media_base():
    base = getattr(settings, 'PUBLIC_MEDIA_BASE_URL', '')
    return base.rstrip('/') if base else ''


def _build_public_asset_url(value):
    cleaned = _clean_str(value)
    if not cleaned:
        return ''
    lowered = cleaned.lower()
    if lowered.startswith('http://') or lowered.startswith('https://'):
        return cleaned
    base = _public_media_base()
    if base:
        return f"{base}/{cleaned.lstrip('/')}"
    return cleaned


def _split_image_refs(raw_value):
    if not raw_value:
        return []
    if isinstance(raw_value, (list, tuple)):
        candidates = raw_value
    else:
        candidates = re.split(r'[\n,|]+', str(raw_value))
    normalized = []
    for candidate in candidates:
        url = _build_public_asset_url(candidate)
        if url:
            normalized.append(url)
    return normalized


def _save_uploaded_inventory_files(inventory, uploads):
    if not uploads:
        return 0
    created = 0
    for upload in uploads:
        url = None
        if getattr(settings, 'S3_UPLOAD_BUCKET', None):
            try:
                _, url = upload_fileobj(upload, folder='inventory')
            except S3UploadError as exc:
                raise ValueError(f"Unable to upload {upload.name}: {exc}") from exc
        else:
            # Developer fallback: store locally if S3 is not configured.
            file_path = default_storage.save(upload.name, upload)
            url = f"{settings.MEDIA_URL.rstrip('/')}/{file_path}"
        inventory_image = InventoryImage.objects.create(
            inventory=inventory,
            image_url=url,
        )
        _mirror_inventory_image_to_storefront(inventory_image)
        created += 1
    return created


def _attach_image_urls(inventory, urls):
    if not urls:
        return 0
    existing = set(
        inventory.inventoryimage_set.filter(image_url__in=urls).values_list('image_url', flat=True)
    )
    created = 0
    for url in urls:
        if url in existing:
            continue
        inv_image = InventoryImage.objects.create(inventory=inventory, image_url=url)
        _mirror_inventory_image_to_storefront(inv_image)
        created += 1
        existing.add(url)
    return created


def _mirror_inventory_image_to_storefront(inventory_image):
    product = inventory_image.inventory.product
    storefront_listing = getattr(product, 'storefront_listing', None)
    if not storefront_listing or not inventory_image.image_url:
        return
    if storefront_listing.images.filter(image_url=inventory_image.image_url).exists():
        return
    display_order = storefront_listing.images.count()
    StorefrontProductImage.objects.create(
        product=storefront_listing,
        image_url=inventory_image.image_url,
        display_order=display_order,
        alt_text=product.name,
    )


def _resolve_category(category_name):
    normalized = _clean_str(category_name)
    if not normalized:
        return None

    tenant_id = get_current_tenant()
    base_qs = Category._base_manager
    lookup = {'name__iexact': normalized}

    category = None
    if tenant_id is not None:
        category = base_qs.filter(tenant_id=tenant_id, **lookup).first()

    if category is None:
        category = base_qs.filter(tenant_id__isnull=True, **lookup).first()

    if category:
        return category

    category = Category(name=normalized)
    category.save()
    return category


DEFAULT_FALLBACK_LOCATION = "Warehouse"


def _prefetch_categories(category_names):
    """
    Return dict mapping lowercase category name -> Category instance for the current tenant (plus shared NULL tenant).
    """
    normalized = {name.lower(): name for name in category_names if name}
    if not normalized:
        return {}

    tenant_id = get_current_tenant()
    qs = Category._base_manager.annotate(lower_name=Lower('name'))

    if tenant_id is not None:
        qs = qs.filter(Q(tenant_id=tenant_id) | Q(tenant_id__isnull=True))
    else:
        qs = qs.filter(tenant_id__isnull=True)

    qs = qs.filter(lower_name__in=normalized.keys())
    return {obj.lower_name: obj for obj in qs}


def _prefetch_products(product_names):
    """
    Return dict mapping lowercase product name -> Product instance for current tenant.
    """
    normalized = {name.lower(): name for name in product_names if name}
    if not normalized:
        return {}

    qs = Product.objects.annotate(lower_name=Lower('name')).filter(lower_name__in=normalized.keys())
    return {obj.lower_name: obj for obj in qs}


def _process_inventory_rows(rows):
    """
    Given a list of normalized row payloads, upsert products/inventory while minimizing queries.
    """
    if not rows:
        return 0

    category_names = {row['category_name'] for row in rows if row['category_name']}
    product_names = {row['product_name'] for row in rows}

    category_cache = _prefetch_categories(category_names)
    product_cache = _prefetch_products(product_names)
    inventory_cache = {}

    imported = 0
    for row in rows:
        category_obj = None
        if row['category_name']:
            key = row['category_name'].lower()
            category_obj = category_cache.get(key)
            if category_obj is None:
                category_obj = Category(name=row['category_name'])
                category_obj.save()
                category_cache[key] = category_obj

        prod_key = row['product_name'].lower()
        product = product_cache.get(prod_key)
        if product is None:
            product = Product.objects.create(
                name=row['product_name'],
                model=row['model_effective'],
                cost=row['cost'],
                price=row['price'],
                description=row['description'],
                category=category_obj,
            )
            product_cache[prod_key] = product
        else:
            updated_fields = []
            if product.cost != row['cost']:
                product.cost = row['cost']
                updated_fields.append('cost')
            if product.price != row['price']:
                product.price = row['price']
                updated_fields.append('price')
            if row['description'] and product.description != row['description']:
                product.description = row['description']
                updated_fields.append('description')
            if category_obj and product.category_id != category_obj.id:
                product.category = category_obj
                updated_fields.append('category')
            if row['model']:
                if product.model != row['model_effective']:
                    product.model = row['model_effective']
                    updated_fields.append('model')
            elif not getattr(product, 'model', None):
                product.model = row['model_effective']
                updated_fields.append('model')

            if updated_fields:
                product.save(update_fields=updated_fields)

        inv_key = (product.id, row['location'].lower())
        inventory = inventory_cache.get(inv_key)
        if inventory is None:
            inventory = Inventory.objects.filter(product=product, location=row['location']).first()
            inventory_cache[inv_key] = inventory

        if inventory:
            update_fields = []
            if inventory.location != row['location']:
                inventory.location = row['location']
                update_fields.append('location')
            if inventory.quantity != row['quantity']:
                inventory.quantity = row['quantity']
                update_fields.append('quantity')
            if update_fields:
                inventory.save(update_fields=update_fields)
        else:
            inventory = Inventory.objects.create(
                product=product,
                location=row['location'],
                quantity=row['quantity'],
            )
            inventory_cache[inv_key] = inventory

        if row.get('image_refs'):
            _attach_image_urls(inventory, row['image_refs'])

        imported += 1

    return imported

@login_required(login_url='/authentication/login/')
@role_required(allowed_roles=['admin', 'manager'])
def index(request):
    context = {
        'message': 'Hello World!'
    }
    return render(request, 'inventory/page/product_index_page.html', context)


@login_required(login_url='/authentication/login/')
@permission_required('inventory.view_product', login_url='/authentication/login/',raise_exception=True)
def create_or_edit_product_info(request, product_id=None):
    product_instance = get_object_or_404(Product, pk=product_id) if product_id else None
    inventory_instance = None

    if product_instance:
        try:
            inventory_instance = Inventory.objects.get(product=product_instance)
        except Inventory.DoesNotExist:
            pass

    if request.method == 'POST':
        product_form = ProductForm(request.POST, instance=product_instance)
        inventory_form = InventoryForm(request.POST, instance=inventory_instance)
        image_form = InventoryImageUploadForm(request.POST, request.FILES)

        if product_form.is_valid() and inventory_form.is_valid() and image_form.is_valid():
            try:
                product = product_form.save()
                inventory = inventory_form.save(commit=False)
                inventory.product = product
                inventory.save()

                uploads = image_form.cleaned_data.get('images')
                manual_urls = _split_image_refs(image_form.cleaned_data.get('image_urls'))
                uploaded_count = 0
                url_count = 0
                if uploads:
                    uploaded_count = _save_uploaded_inventory_files(inventory, uploads)
                if manual_urls:
                    url_count = _attach_image_urls(inventory, manual_urls)

                if uploaded_count or url_count:
                    messages.success(
                        request,
                        f"Product saved with {uploaded_count} upload(s) and {url_count} linked image(s)."
                    )
                else:
                    messages.success(request, 'Product saved successfully!')
                return redirect('inventory:item_list')
            except (ValueError, S3UploadError) as e:
                messages.error(request, f'Image upload failed: {str(e)}')
            except Exception as e:
                messages.error(request, f'Error saving product: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        product_form = ProductForm(instance=product_instance)
        inventory_form = InventoryForm(instance=inventory_instance)
        image_form = InventoryImageUploadForm()

    context = {
        'product_form': product_form,
        'inventory_form': inventory_form,
        'image_form': image_form,
        'product_id': product_id,
        'is_edit': product_id is not None
    }

    return render(request, 'inventory/page/product_upload_page.html', context)

@login_required(login_url='/authentication/login/')
def item_list_view(request):
    products = Product.objects.all()
    inventory_data = Inventory.objects.all()

    product_images = {}
    image_qs = (
        InventoryImage.objects.filter(inventory__product__in=products)
        .order_by('inventory__product_id', 'id')
        .values('inventory__product_id', 'image_url')
    )
    for entry in image_qs:
        pid = entry['inventory__product_id']
        if pid not in product_images and entry['image_url']:
            product_images[pid] = entry['image_url']

    # Create a mapping of product to stock
    product_inventory = {}
    for inventory in inventory_data:
        product_id = inventory.product.id
        if product_id in product_inventory:
            product_inventory[product_id] += inventory.quantity
        else:
            product_inventory[product_id] = inventory.quantity

    # Add stock data to products
    product_list = [
        {
            'id': product.id,
            'name': product.name,
            'category': product.category,
            'cost': product.cost,
            'price': product.price,
            'stock': product_inventory.get(product.id, 0),
            'image_url': product_images.get(product.id),
        }
        for product in products
    ]

    # Compute total inventory value = sum(price * stock)
    total_value = sum(
        Decimal(str(item['price'])) * Decimal(item['stock'])
        for item in product_list
    ) if product_list else Decimal('0.00')

    low_stock_count = sum(1 for item in product_list if (item.get('stock') or 0) <= 5)
    context = {
        'products': product_list,
        'total_items': products.count(),
        'total_value': total_value,
        'low_stock_count': low_stock_count,
    }
    return render(request, 'inventory/page/item_list_page.html', context)


@login_required(login_url='/authentication/login/')
def product_delete_view(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    product.delete()
    messages.success(request, 'Product Deleted Successfully!')
    return redirect('inventory:item_list')


@login_required(login_url='/authentication/login/')
def upload_images(request, product_id):
    InventoryImageFormSet = inlineformset_factory(Inventory, InventoryImage,
                                                  form=InventoryImageForm, extra=1)
    inventory_instance = get_object_or_404(Inventory, pk=product_id)

    if request.method == 'POST':
        formset = InventoryImageFormSet(request.POST, request.FILES, instance=inventory_instance)
        if formset.is_valid():

            for form in formset.forms:
                if form.cleaned_data and form.cleaned_data.get('image'):
                    try:
                        _save_uploaded_inventory_files(inventory_instance, form.cleaned_data['image'])
                    except ValueError as exc:
                        messages.error(request, str(exc))
                        break
            else:
                messages.success(request, 'Images uploaded successfully!')
                return redirect('inventory:edit_product', product_id=product_id)
        else:
            print("Formset Errors:", formset.errors)

    else:
        formset = InventoryImageFormSet()

    return render(request, 'inventory/page/image_upload.html', {'formset': formset})

@login_required(login_url='/authentication/login/')
def export_inventory(request):
    # Create a workbook and select active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Add header row
    headers = [
        "Product Name", "Model", "Category", "Cost", "Price",
        "Description", "Location", "Quantity",
    ]
    ws.append(headers)

    # Query inventory and product data
    inventory_items = Inventory.objects.select_related('product').all()

    # Populate rows
    for item in inventory_items:
        ws.append([
            item.product.name,
            item.product.model or "",
            item.product.category.name if item.product.category else "",
            item.product.cost,
            item.product.price,
            item.product.description,
            item.location,
            item.quantity,
        ])

    # Create response with content type for Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=inventory.xlsx'
    wb.save(response)
    return response


@login_required(login_url='/authentication/login/')
def import_inventory(request):
    if request.method != "POST" or 'file' not in request.FILES:
        if _wants_json(request):
            return JsonResponse({'status': 'error', 'message': 'No file selected or invalid request.'}, status=400)
        messages.error(request, "No file selected or invalid request.")
        return redirect('inventory:item_list')

    file = request.FILES['file']
    file_name = (file.name or '').lower()
    wants_json = _wants_json(request)

    headers = [
        "product name", "model", "category", "cost", "price", "description", "location", "quantity", "image urls"
    ]
    required_headers = {"product name", "cost", "price", "location", "quantity"}
    seen_header_map = {}
    imported_lines = 0
    errors = []
    row_payloads = []

    try:
        with transaction.atomic():
            if file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active

                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                idx_map = {_normalize_header(h): i for i, h in enumerate(header_row or []) if _normalize_header(h)}
                seen_header_map = idx_map

                missing_required = sorted([h for h in required_headers if h not in idx_map])
                if missing_required:
                    msg = f"Missing required columns: {', '.join(missing_required)}"
                    raise ValueError(msg)

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row:
                        continue

                    raw_product = row[idx_map.get('product name')] if idx_map.get('product name') is not None else None
                    raw_model = row[idx_map.get('model')] if idx_map.get('model') is not None else None
                    raw_category = row[idx_map.get('category')] if idx_map.get('category') is not None else None
                    raw_description = row[idx_map.get('description')] if idx_map.get('description') is not None else None
                    raw_location = row[idx_map.get('location')] if idx_map.get('location') is not None else None
                    cost_val = row[idx_map.get('cost')] if idx_map.get('cost') is not None else None
                    price_val = row[idx_map.get('price')] if idx_map.get('price') is not None else None
                    quantity_val = row[idx_map.get('quantity')] if idx_map.get('quantity') is not None else None
                    raw_images = row[idx_map.get('image urls')] if idx_map.get('image urls') is not None else None

                    product_name = _clean_str(raw_product)
                    model_clean = _clean_str(raw_model)
                    model_effective = model_clean or product_name
                    category_name = _clean_str(raw_category)
                    description = _clean_str(raw_description)
                    location = _clean_str(raw_location) or DEFAULT_FALLBACK_LOCATION

                    cost = _parse_decimal(cost_val)
                    price = _parse_decimal(price_val)
                    quantity = _parse_int(quantity_val)
                    if quantity is not None and quantity < 0:
                        quantity = 0
                    image_refs = _split_image_refs(raw_images)

                    missing_bits = []
                    if not product_name:
                        missing_bits.append('product name')
                    if cost is None:
                        missing_bits.append('cost')
                    if price is None:
                        missing_bits.append('price')
                    if quantity is None:
                        missing_bits.append('quantity')
                    if missing_bits:
                        errors.append(
                            f"Row {row_idx}: Missing {', '.join(missing_bits)}. "
                            f"Raw values -> product={raw_product!r}, model={raw_model!r}, "
                            f"category={raw_category!r}, location={raw_location!r}, "
                            f"cost={cost_val!r}, price={price_val!r}, quantity={quantity_val!r}"
                        )
                        continue

                    row_payloads.append({
                        'product_name': product_name,
                        'model': model_clean,
                        'model_effective': model_effective,
                        'category_name': category_name,
                        'description': description[:200] if description else '',
                        'location': location,
                        'cost': cost,
                        'price': price,
                        'quantity': quantity,
                        'image_refs': image_refs,
                    })

            elif file_name.endswith('.csv'):
                # Minimal CSV support: use the same headers as the xlsx template.
                decoded = file.read().decode('utf-8-sig')
                reader = csv.DictReader(decoded.splitlines())
                if not reader.fieldnames:
                    raise ValueError("CSV is missing headers.")

                idx_map = {_normalize_header(h): h for h in reader.fieldnames}
                missing_required = sorted([h for h in required_headers if h not in idx_map])
                if missing_required:
                    msg = f"Missing required columns: {', '.join(missing_required)}"
                    raise ValueError(msg)

                for row_idx, row in enumerate(reader, start=2):
                    raw_product = row.get(idx_map.get('product name'))
                    raw_model = row.get(idx_map.get('model'))
                    raw_location = row.get(idx_map.get('location'))
                    raw_category = row.get(idx_map.get('category'))
                    raw_description = row.get(idx_map.get('description'))
                    raw_cost = row.get(idx_map.get('cost'))
                    raw_price = row.get(idx_map.get('price'))
                    raw_quantity = row.get(idx_map.get('quantity'))
                    raw_images = row.get(idx_map.get('image urls'))

                    product_name = _clean_str(raw_product)
                    model_clean = _clean_str(raw_model)
                    location = _clean_str(raw_location) or DEFAULT_FALLBACK_LOCATION
                    category_name = _clean_str(raw_category)
                    description = _clean_str(raw_description)
                    cost = _parse_decimal(raw_cost)
                    price = _parse_decimal(raw_price)
                    quantity = _parse_int(raw_quantity)
                    if quantity is not None and quantity < 0:
                        quantity = 0
                    image_refs = _split_image_refs(raw_images)

                    missing_bits = []
                    if not product_name:
                        missing_bits.append('product name')
                    if cost is None:
                        missing_bits.append('cost')
                    if price is None:
                        missing_bits.append('price')
                    if quantity is None:
                        missing_bits.append('quantity')

                    if missing_bits:
                        errors.append(
                            f"Row {row_idx}: Missing {', '.join(missing_bits)}. "
                            f"Raw values -> product={raw_product!r}, model={raw_model!r}, "
                            f"category={raw_category!r}, location={raw_location!r}, "
                            f"cost={raw_cost!r}, price={raw_price!r}, quantity={raw_quantity!r}"
                        )
                        continue

                    row_payloads.append({
                        'product_name': product_name,
                        'model': model_clean,
                        'model_effective': (model_clean or product_name),
                        'category_name': category_name,
                        'description': description[:200] if description else '',
                        'location': location,
                        'cost': cost,
                        'price': price,
                        'quantity': quantity,
                        'image_refs': image_refs,
                    })
            else:
                raise ValueError("Unsupported file type. Upload a .xlsx or .csv file.")

            imported_lines = _process_inventory_rows(row_payloads)

    except Exception as e:
        if wants_json:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
        messages.error(request, f"Error during import: {str(e)}")
        return redirect('inventory:item_list')

    if wants_json:
        return JsonResponse({
            'status': 'ok',
            'imported_lines': imported_lines,
            'error_count': len(errors),
            'errors': errors[:10],
        })

    if errors:
        sample = " | ".join(errors[:3])
        messages.warning(
            request,
            f"Imported with {len(errors)} row errors. Successfully imported {imported_lines} lines. Sample issues: {sample}"
        )
    else:
        messages.success(request, f"Inventory imported successfully. Imported {imported_lines} lines.")
    return redirect('inventory:item_list')


@login_required(login_url='/authentication/login/')
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    headers = ["Product Name", "Model", "Category", "Cost", "Price", "Description", "Location", "Quantity", "Image URLs"]
    ws.append(headers)

    # Optional example row (kept blank-ish to reduce mistakes).
    ws.append([
        "Example Item", "Example Model", "Default Category",
        10.00, 15.00, "", "Warehouse A", 5,
        "case-front.jpg|https://cdn.example.com/cases/case-back.jpg"
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=inventory_template.xlsx'
    wb.save(response)
    return response


@login_required(login_url='/authentication/login/')
def category_list(request):
    categories = Category.objects.all()
    context = {
        'categories': categories,
    }
    return render(request, 'inventory/page/category_list_page.html', context)


@login_required(login_url='/authentication/login/')
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully!')
            return redirect('inventory:category_list')
    else:
        form = CategoryForm()

    return render(request, 'inventory/page/category_form.html', {
        'form': form,
        'is_add': True
    })


@login_required(login_url='/authentication/login/')
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully!')
            return redirect('inventory:category_list')
    else:
        form = CategoryForm(instance=category)

    return render(request, 'inventory/page/category_form.html', {
        'form': form,
        'category': category,
        'is_add': False
    })


@login_required(login_url='/authentication/login/')
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        try:
            category.delete()
            messages.success(request, 'Category deleted successfully!')
        except Exception as e:
            messages.error(request, 'Cannot delete category. It has associated products.')
    return redirect('inventory:category_list')
